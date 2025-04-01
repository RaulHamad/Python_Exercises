
import openpyxl
from openpyxl.styles import  Font
from openpyxl.chart import BarChart

#fig =plt.subplot()
#fig.plot([1,2,3],[1,2,3])
#plt.show()



wb = openpyxl.Workbook() #cria tabela
ws = wb.active # planilha inicial "sheet" ativa
ws.title='graphic' #mudei nome da planilha sheet para grafic2
#ws1 = wb.create_sheet('init',0)
planilha_graphic = wb['graphic']


#ws = wb['init']
print(wb.active)
#ws.title = 'toto'
#for i in wb.sheetnames:
#    print(i)
#folha = 'grafic'
#wb.create_sheet(folha)#criando nova folha para ficheiro
print(wb.sheetnames)
#wb['graphic'].insert_rows(1)
planilha_graphic["A1"] = "Date"
planilha_graphic["B1"] = "Expense"
planilha_graphic["C1"] = "Profit"
planilha_graphic["D1"] = "Total"

bold_font = Font(bold=True)


for cell in planilha_graphic["A1":"D1"]:
    for row in cell:
        row.font = bold_font




def paises_e_medalhas(date,profit,expense,total):
    """
    Função que adiciona cada país e suas medalhas em uma lista temporaria, adicionando a lista_olimpiada
    pais: nome do país a adicionar
    ouro: quantidade de medalhas de ouro do país
    prata: quantidade de medalhas de prata do país
    bronze: quantidade de medalhas de bronze do país
    """
    lista_temp = [date,profit,expense,total]
    planilha_graphic.append(lista_temp)
    
    return f'data : {date},\n,\nlucro: {profit},\ndebito: {expense},\n total: {total}'
  
list_items = [["USA", 46, 12, 5],["China", 38, 20, 7],["UK", 29, 7, 7],["Russia", 22, 10, 9]]   
for i in list_items:
    paises_e_medalhas(i[0],i[1],i[2],(i[1]-i[2]))
   

#Criando grafico de barras
chart = BarChart()
chart.type = "col"
chart.title = "Status"
chart.y_axis.title = 'Profit'
chart.x_axis.title = 'Date'
chart.legend = None
chart.varyColors = True # colocar cores diferentes nas barras

#Criando as barras e lançando dados no grafico
data = openpyxl.chart.Reference(planilha_graphic, min_col=4, min_row=2, max_col=5, max_row=7)
categs = openpyxl.chart.Reference(planilha_graphic, min_col=1, min_row=2, max_row=7)
chart.add_data(data)
chart.set_categories(categs)

planilha_graphic.add_chart(chart, "G1")

wb.save('grafico.xlsx')

import tkinter as tk
from tkinter import ttk

import matplotlib
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import pandas as pd


# DataFrame para ejemplo
df = pd.DataFrame({"Foo": (1, 5, 8, 7, 5, 1, 5),
                   "Bar": (4.25, 5, 6, 1.3, 1, 2.6, 3.7)}
                  )

# Creamos una figura y añadimos algunas gráficas en subplots
fig1 = Figure(figsize=(5, 5), dpi=100)
ax1 = fig1.add_subplot(311)
ax2 = fig1.add_subplot(312)
ax3 = fig1.add_subplot(313)
df.plot(ax=ax1)
df.plot.bar(ax=ax2)
df.plot.area(ax=ax3)


# Ventana principal
root = tk.Tk()

# Canvas  y barras de desplazamiento
canvas = tk.Canvas(root, borderwidth=0, background="#ffffff")
hsb = tk.Scrollbar(root, orient="horizontal", command=canvas.xview)                 
vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
canvas.configure(xscrollcommand=hsb.set)
canvas.configure(yscrollcommand=vsb.set)
hsb.pack(side="bottom", fill="x")
vsb.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)

# Vamos a usar un ttk.Labelframe para contener FigureCanvasTkAgg
plot_frame = ttk.Labelframe(canvas, text='Plots')
canvas.create_window((4,4), window=plot_frame, anchor="nw", tags="plot_frame")
plot_frame.bind("<Configure>",
                lambda event: canvas.configure(scrollregion=canvas.bbox("all"))
                )

# Creamos una instancia de FigureCanvas que renderizará la figura
canvas1 = FigureCanvasTkAgg(fig1, master=plot_frame)
canvas1.draw()
canvas1.get_tk_widget().grid(row=0, column=0)

root.mainloop()